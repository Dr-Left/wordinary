""" wordExtract.py - Extract words from texts in bin and return a dictionary
with keys of words and values of the times the word appears.

p.s. Compatible with documents ends with .txt, .docx


Typical usage example:

wordList = wordExtract.extract()

arguments: <openPath1> <savePath> [basisPath]

"""

import codecs
import glob
import logging
import os
import re
import shelve
import sys
import time

import docx
import openpyxl
from openpyxl.styles import Font

import currentTime
import shelfer  # my own module shelfer.py

basisPath = ""


def output(fname, wordList):
    # generate the shelf of the dictionary in .\data
    shelfer.work(basisPath)  # if the GUI not send it, it remains blank
    # create an excel file
    wb = openpyxl.Workbook()  # workbook
    ws = wb.active  # worksheet
    # initial the headline
    ws["A1"] = "序号"
    ws["B1"] = "单词"
    ws["C1"] = "释意"
    ws["D1"] = "词频"
    # output the words and the numbers
    line = 1
    shelfFile = shelve.open(".\\data\dict")
    if shelfFile.keys() == []:
        raise Exception("Run shelfer.py first!")
    for word, num in wordList.items():
        try:
            # Another mode:get from the internet: translation = trans.getTranslation(word)
            if word not in shelfFile.keys():
                continue
            translation = shelfFile[word]
            line += 1
            ws["A" + str(line)] = line - 1
            ws["B" + str(line)] = word
            ws["C" + str(line)] = translation
            ws["D" + str(line)] = num
        except Exception as exc:
            print("Translation error: %s" % (exc))

    # set the column width
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 17
    ws.column_dimensions["C"].width = 90
    ws.column_dimensions["D"].width = 7
    # set the font
    myFont = Font(name="Times New Roman", size=11)
    for i in range(len(wordList) + 1):
        for j in range(4):
            ws.cell(i + 1, j + 1).font = myFont  # cell 是从1开始
    # save
    wb.save(fname)
    print(os.path.abspath(fname))
    print(line - 1)
    return 0


def cutUp(text):
    text = text.lower()
    # not an English alphabet
    myRegex = re.compile(r"[^a-zA-Z]+")  # this should be plus but not star
    text = myRegex.sub(" ", text)

    # print(myRegex.findall(text))
    # print('\n\n\n\n' + text)

    return text


def extract(openPath, savePath):
    currentDirectory = os.getcwd()
    words = {}
    # print(openPath)
    # for each text, collect the words and the corresponding appearing times
    dr = openPath
    if dr.endswith(".txt"):
        try:
            myFile = open(dr)
        except Exception as exc:
            print("cannot open directory.")
        try:
            cont = myFile.read()
        except Exception as exc:
            myFile = codecs.open(dr, mode="r", encoding="utf-8")
            cont = myFile.read()
        cont = cutUp(cont)
        wordList = cont.split()
        for wd in wordList:
            words.setdefault(wd, 0)
            words[wd] += 1
    elif dr.endswith(".doc") or dr.endswith(".docx"):
        try:
            myFile = docx.Document(dr)
        except Exception as exc:
            print("cannot open directory.")
        cont = ""
        for paragraph in myFile.paragraphs:
            cont += paragraph.text
        cont = cutUp(cont)
        wordList = cont.split()
        for wd in wordList:
            words.setdefault(wd, 0)
            words[wd] += 1

    # words = sorted(words.items())
    # print('词语提取成功')
    # print(words)
    # time.sleep(0.5)
    # print(words)
    return words  #    return a dictionary of words


if __name__ == "__main__":
    if len(sys.argv) < 3:
        sys.exit(-1)
    openPath = sys.argv[1]
    savePath = sys.argv[2]
    if len(sys.argv) > 3:
        basisPath = sys.argv[3]
    words = extract(openPath, savePath)
    output(savePath, words)
