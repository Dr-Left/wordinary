'''outputToExcel.py - output the word result to an excel workbook.

'''

import wordExtract
import openpyxl
import currentTime
import os
import shelve
import shelfer # my own module shelfer.py
import logging

logging.basicConfig(level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s')
logging.disable(logging.CRITICAL)

def output():
    # generate the shelf of the dictionary in .\data
    shelfer.work()
    
   
    
    #do the extraction work

    wordList = wordExtract.extract()
    logging.debug(wordList)


    # create an excel file

    wb = openpyxl.Workbook()    # workbook
    ws = wb.active   # worksheet
    # initial the headline
    ws['A1'] = '序号'
    ws['B1'] = '单词'
    ws['C1'] = '释意'
    ws['D1'] = '词频'
    
    # output the words and the numbers
    
    line = 1

    shelfFile = shelve.open('.\\data\dict')
    if shelfFile.keys() == []:
        raise Exception('Run shelfer.py first!')

    for word, num in wordList:
        try:
            #Another mode:get from the internet: translation = trans.getTranslation(word)

            if word not in shelfFile.keys():
                continue
            
            translation = shelfFile[word]
            
            line += 1
            ws['A' + str(line)] = line - 1
            ws['B' + str(line)] = word
            ws['C' + str(line)] = translation
            ws['D' + str(line)] = num
        except Exception as exc:
            print('Translation error: %s' %(exc))
            
    # set the column width
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['C'].width = 90
    ws.column_dimensions['D'].width = 7
    
    # file directory work
    cT = currentTime.getTime()
    binPath = '.\\生成的文件\\'
    if not os.path.exists(binPath):
        os.makedirs(binPath)
    fname = binPath + '词频表 ' + cT + '.xlsx'
    
    #save
    wb.save(fname)
    print('词频表生成成功，位于\t' + os.path.abspath(fname))


if __name__=='__main__':
    print('main')
    output()
